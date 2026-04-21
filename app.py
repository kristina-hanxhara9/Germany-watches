import os
import io
import re
import sqlite3
import xml.etree.ElementTree as ET

import requests
from flask import Flask, request, jsonify, render_template, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)

# Two German databases
DB_FLAT = os.path.join(os.path.dirname(__file__), "openregister.db")       # flat, 5.3M
DB_NORM = os.path.join(os.path.dirname(__file__), "handelsregister.db")    # normalized, 2.2M with objectives
SOAP_URL = "https://justizonline.gv.at/jop/api/at.gv.justiz.fbw/ws/fbw.wsdl"


# =====================================================================
#  NACE keyword classifier (German business purpose -> NACE division)
# =====================================================================

NACE_KEYWORDS = {
    "01": ["landwirtschaft","ackerbau","viehzucht","tierhaltung","pflanzenbau","agrar"],
    "02": ["forstwirtschaft","holzeinschlag","waldwirtschaft"],
    "03": ["fischerei","aquakultur","fischzucht"],
    "05": ["kohlebergbau","kohlenbergbau","braunkohle","steinkohle"],
    "06": ["erdöl","erdgas","petroleum"],
    "07": ["erzbergbau","eisenerz","metallerz"],
    "08": ["steinbruch","kies","sand","ton","kaolin"],
    "10": ["nahrungsmittel","lebensmittel","fleisch","fisch","obst","gemüse","milch","käse","backwaren","bäckerei","metzgerei","schlachtung"],
    "11": ["getränke","brauerei","bier","wein","spirituosen","saft","mineralwasser"],
    "13": ["textil","weberei","spinnerei","stoff"],
    "14": ["bekleidung","kleidung","mode","schneiderei","konfektion"],
    "15": ["leder","schuhe","schuhwerk","lederwaren","gerberei"],
    "16": ["holzverarbeitung","sägewerk","tischlerei","möbel","holzwaren","zimmerei"],
    "17": ["papier","pappe","zellstoff","papierherstellung"],
    "18": ["druckerei","druck","vervielfältigung","druckerzeugnis"],
    "19": ["kokerei","mineralöl","raffinerie"],
    "20": ["chemie","chemisch","chemikalien","farben","lacke","klebstoff","düngemittel","pestizid"],
    "21": ["pharma","arzneimittel","medikament","pharmazeutisch"],
    "22": ["gummi","kunststoff","plastik","reifen"],
    "23": ["glas","keramik","zement","beton","ziegel","fliesen","stein"],
    "24": ["metall","stahl","eisen","aluminium","kupfer","gießerei","hütte"],
    "25": ["metallerzeugnisse","metallbau","schlosserei","metallverarbeitung","werkzeug","schrauben"],
    "26": ["elektronik","halbleiter","computer","datenverarbeitung","optik","messtechnik","sensorik"],
    "27": ["elektrisch","elektrotechnik","elektromotor","kabel","batterie","akku","beleuchtung","lampe"],
    "28": ["maschinenbau","maschine","anlage","turbine","pumpe","kompressor","werkzeugmaschine"],
    "29": ["automobil","kraftfahrzeug","fahrzeug","kfz","auto","karosserie","kraftwagen"],
    "30": ["schiffbau","flugzeug","schienenfahrzeug","lokomotive","waggon"],
    "31": ["möbel","einrichtung","küchenmöbel","büromöbel","matratze"],
    "32": ["schmuck","musikinstrument","sportartikel","spielwaren","medizintechnik","zahntechnik","dental"],
    "33": ["reparatur","wartung","instandhaltung","installation"],
    "35": ["energie","strom","gas","elektrizität","heizung","fernwärme","solar","wind","photovoltaik","erneuerbare"],
    "36": ["wasserversorgung","trinkwasser","wasseraufbereitung"],
    "37": ["abwasser","kanalisation","kläranlage"],
    "38": ["abfall","entsorgung","recycling","müll","wertstoff","altpapier","schrott"],
    "39": ["sanierung","dekontamination","altlasten","umweltsanierung"],
    "41": ["hochbau","wohnungsbau","bauträger","bauunternehmen","baugesellschaft"],
    "42": ["tiefbau","straßenbau","brückenbau","tunnelbau","leitungsbau"],
    "43": ["bauinstallation","elektroinstallation","sanitär","heizungsinstallation","malerei","trockenbau","dachdecker","gerüstbau","abbruch"],
    "45": ["autohandel","kfz-handel","autowerkstatt","kfz-werkstatt","gebrauchtwagen"],
    "46": ["großhandel","wholesale","import","export","handelsvertretung","vermittlung"],
    "47": ["einzelhandel","laden","geschäft","verkauf","retail","supermarkt","kaufhaus","versandhandel","onlinehandel","e-commerce"],
    "49": ["transport","spedition","logistik","güterverkehr","personenverkehr","taxi","bus","eisenbahn","pipeline"],
    "50": ["schifffahrt","seeverkehr","binnenschifffahrt"],
    "51": ["luftfahrt","flugverkehr","luftverkehr","airline"],
    "52": ["lagerei","lager","umschlag","parkhaus","parkplatz"],
    "53": ["post","kurier","express","brief","paket"],
    "55": ["hotel","beherbergung","pension","gasthof","ferienwohnung","camping","jugendherberge"],
    "56": ["gastronomie","restaurant","gaststätte","imbiss","catering","kantine","café","bar","kneipe"],
    "58": ["verlag","buch","zeitung","zeitschrift","software","publishing"],
    "59": ["film","video","fernsehen","kino","musik","tonträger","produktion"],
    "60": ["rundfunk","radio","fernsehprogramm","broadcasting"],
    "61": ["telekommunikation","telefon","mobilfunk","internet","provider","netzwerk"],
    "62": ["software","programmierung","it-beratung","edv","informatik","datenbank","app","webentwicklung","it-dienstleistung","systemhaus","digitalisierung"],
    "63": ["datenverarbeitung","hosting","portal","suchmaschine","informationsdienst"],
    "64": ["bank","kredit","finanzdienstleistung","investmentfonds","beteiligung","holding","vermögensverwaltung","treuhand","kapitalanlage"],
    "65": ["versicherung","rückversicherung","pensionskasse","vorsorge"],
    "66": ["finanzmakler","wertpapier","börse","fondsverwaltung","versicherungsmakler"],
    "68": ["immobilien","grundstück","wohnung","hausverwaltung","maklertätigkeit","immobilienverwaltung","vermietung","grundbesitz"],
    "69": ["rechtsanwalt","anwalt","steuerberater","wirtschaftsprüfer","notar","buchführung","buchhaltung","rechtsberatung"],
    "70": ["unternehmensberatung","managementberatung","consulting","beratung","geschäftsführung"],
    "71": ["architekt","ingenieur","planung","prüfung","gutachten","vermessung","statik","bauplanung"],
    "72": ["forschung","entwicklung","wissenschaft","labor","biotechnologie"],
    "73": ["werbung","marketing","marktforschung","kommunikation","pr","public relations","medien"],
    "74": ["design","fotografie","übersetzung","dolmetscher","sachverständige"],
    "75": ["tierarzt","veterinär","tierklinik","tiermedizin"],
    "77": ["vermietung","leasing","verleih","mietwagen"],
    "78": ["personalvermittlung","zeitarbeit","arbeitnehmerüberlassung","personaldienstleistung"],
    "79": ["reisebüro","reiseveranstalter","touristik","tourismus","reise"],
    "80": ["sicherheitsdienst","wachdienst","bewachung","detektei","sicherheitsberatung"],
    "81": ["gebäudereinigung","reinigung","hausmeister","facility","garten","landschaftsbau","gebäudemanagement"],
    "82": ["callcenter","sekretariat","bürodienstleistung","inkasso","messe","kongress","veranstaltung","event"],
    "85": ["schule","bildung","weiterbildung","ausbildung","unterricht","nachhilfe","seminar","coaching","training","akademie","fahrschule"],
    "86": ["arzt","klinik","krankenhaus","praxis","gesundheit","pflege","therapie","physiotherapie","zahnarzt","heilpraktiker","apotheke"],
    "87": ["pflegeheim","altenheim","seniorenheim","betreutes wohnen"],
    "88": ["sozial","jugendhilfe","behindertenhilfe","kindertagesstätte","kita"],
    "90": ["kunst","theater","musik","unterhaltung","kreativ"],
    "91": ["bibliothek","museum","archiv","denkmal","botanisch","zoo"],
    "92": ["glücksspiel","wetten","lotterie","spielhalle","casino"],
    "93": ["sport","fitnessstudio","fitness","schwimmbad","freizeitpark","vergnügung","bowling"],
    "94": ["verein","verband","gewerkschaft","partei","kirche","stiftung"],
    "95": ["reparatur von computer","reparatur von elektronik","uhrmacher"],
    "96": ["friseur","kosmetik","bestattung","sauna","solarium","wellness","tattoo","piercing","wäscherei"],
}

NACE_LABELS = {
    "01":"Crop & animal production","02":"Forestry & logging","03":"Fishing & aquaculture",
    "05":"Mining of coal","06":"Petroleum & gas extraction","07":"Mining of metal ores",
    "08":"Other mining & quarrying","10":"Food products","11":"Beverages",
    "13":"Textiles","14":"Wearing apparel","15":"Leather products",
    "16":"Wood products","17":"Paper products","18":"Printing",
    "19":"Coke & petroleum","20":"Chemicals","21":"Pharmaceuticals",
    "22":"Rubber & plastic","23":"Glass, cement, ceramics","24":"Basic metals",
    "25":"Fabricated metal products","26":"Computer & electronics","27":"Electrical equipment",
    "28":"Machinery & equipment","29":"Motor vehicles","30":"Other transport equipment",
    "31":"Furniture","32":"Other manufacturing","33":"Repair & installation",
    "35":"Energy supply","36":"Water supply","37":"Sewerage",
    "38":"Waste management","39":"Remediation","41":"Building construction",
    "42":"Civil engineering","43":"Specialised construction","45":"Motor vehicle trade",
    "46":"Wholesale trade","47":"Retail trade","49":"Land transport & logistics",
    "50":"Water transport","51":"Air transport","52":"Warehousing",
    "53":"Postal & courier","55":"Accommodation","56":"Food & beverage service",
    "58":"Publishing","59":"Film & music production","60":"Broadcasting",
    "61":"Telecommunications","62":"IT & software","63":"Information services",
    "64":"Financial services","65":"Insurance","66":"Auxiliary financial services",
    "68":"Real estate","69":"Legal & accounting","70":"Management consultancy",
    "71":"Architecture & engineering","72":"R&D","73":"Advertising & marketing",
    "74":"Other professional services","75":"Veterinary","77":"Rental & leasing",
    "78":"Employment services","79":"Travel & tourism","80":"Security services",
    "81":"Building & landscape services","82":"Office & business support",
    "85":"Education","86":"Health services","87":"Residential care",
    "88":"Social work","90":"Arts & entertainment","91":"Libraries & museums",
    "92":"Gambling","93":"Sports & recreation","94":"Membership organisations",
    "95":"Repair of computers","96":"Personal services",
}


def classify_nace(objective_text):
    """Classify a German business purpose text to NACE division codes."""
    if not objective_text:
        return []
    text = objective_text.lower()
    scores = {}
    for code, keywords in NACE_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                scores[code] = scores.get(code, 0) + 1
    if not scores:
        return []
    max_score = max(scores.values())
    # Return all codes with the max score, plus any with >1 match
    results = []
    for code, score in sorted(scores.items(), key=lambda x: -x[1]):
        if score >= max(1, max_score - 1):
            results.append({"code": code, "label": NACE_LABELS.get(code, ""), "score": score})
    return results[:5]  # Top 5


# =====================================================================
#  GERMANY — normalized SQLite (handelsregister.db)
# =====================================================================

def _get_norm_db():
    conn = sqlite3.connect(DB_NORM)
    conn.row_factory = sqlite3.Row
    return conn


def _get_flat_db():
    conn = sqlite3.connect(DB_FLAT)
    conn.row_factory = sqlite3.Row
    return conn


def de_search(q, page, per_page, nace_code="", use_objective_search=False):
    """Search German companies. If nace_code given, search objectives text."""
    offset = (page - 1) * per_page

    if os.path.exists(DB_NORM):
        return _de_search_norm(q, page, per_page, offset, nace_code, use_objective_search)
    elif os.path.exists(DB_FLAT):
        return _de_search_flat(q, page, per_page, offset)
    else:
        return {"error": "No German database found.", "companies": [], "total": 0}


def _fetch_company_rows(conn, company_ids):
    """Fetch deduplicated company details for a list of companyIds."""
    if not company_ids:
        return []
    results = []
    seen = set()
    for cid in company_ids:
        if cid in seen:
            continue
        seen.add(cid)
        name = conn.execute("SELECT name FROM Names WHERE companyId=? AND isCurrent='True' LIMIT 1", (cid,)).fetchone()
        obj = conn.execute("SELECT objective FROM Objectives WHERE companyId=? LIMIT 1", (cid,)).fetchone()
        addr = conn.execute("SELECT fullAddress, zipAndPlace FROM Addresses WHERE companyId=? AND isCurrent='True' LIMIT 1", (cid,)).fetchone()
        ref = conn.execute("SELECT nativeReferenceNumber, courtName FROM ReferenceNumbers WHERE companyId=? LIMIT 1", (cid,)).fetchone()
        comp = conn.execute("SELECT foundedDate, dissolutionDate FROM Companies WHERE companyId=? LIMIT 1", (cid,)).fetchone()
        cap = conn.execute("SELECT capitalAmount, capitalCurrency FROM Capital WHERE companyId=? AND isCurrent='True' LIMIT 1", (cid,)).fetchone()
        results.append({
            "companyId": cid,
            "name": name["name"] if name else "",
            "objective": obj["objective"] if obj else "",
            "fullAddress": addr["fullAddress"] if addr else "",
            "zipAndPlace": addr["zipAndPlace"] if addr else "",
            "nativeReferenceNumber": ref["nativeReferenceNumber"] if ref else "",
            "courtName": ref["courtName"] if ref else "",
            "foundedDate": comp["foundedDate"] if comp else "",
            "dissolutionDate": comp["dissolutionDate"] if comp else "",
            "capitalAmount": cap["capitalAmount"] if cap else None,
            "capitalCurrency": cap["capitalCurrency"] if cap else "",
        })
    return results


def _de_search_norm(q, page, per_page, offset, nace_code="", use_objective_search=False):
    """Search using the normalized handelsregister.db."""
    conn = _get_norm_db()
    try:
        # If searching by NACE, use the objectives FTS
        if nace_code and nace_code in NACE_KEYWORDS:
            keywords = NACE_KEYWORDS[nace_code][:5]
            # Quote each keyword for FTS5 safety (handles hyphens, spaces)
            fts_query = " OR ".join(f'"{kw}"' for kw in keywords)

            # Step 1: Get matching companyIds from FTS (fast)
            id_sql = "SELECT DISTINCT companyId FROM ObjectivesFts WHERE ObjectivesFts MATCH ?"
            fts_ids = [r[0] for r in conn.execute(id_sql, [fts_query]).fetchall()]

            if not fts_ids:
                return {"companies": [], "total": 0, "page": page, "per_page": per_page}

            # Filter by name if provided
            if q:
                placeholders = ",".join(["?"] * min(len(fts_ids), 5000))
                name_sql = f"SELECT DISTINCT companyId FROM Names WHERE companyId IN ({placeholders}) AND name LIKE ? AND isCurrent = 'True'"
                fts_ids = [r[0] for r in conn.execute(name_sql, fts_ids[:5000] + [f"%{q}%"]).fetchall()]

            total = len(fts_ids)
            page_ids = fts_ids[offset:offset + per_page]

            if not page_ids:
                return {"companies": [], "total": total, "page": page, "per_page": per_page}

            # Step 2: Get details for this page (deduplicated)
            rows = _fetch_company_rows(conn, page_ids)

        elif q:
            # Name search
            count_sql = "SELECT COUNT(*) FROM Names WHERE name LIKE ? AND isCurrent = 'True'"
            total = conn.execute(count_sql, [f"%{q}%"]).fetchone()[0]

            id_sql = "SELECT companyId FROM Names WHERE name LIKE ? AND isCurrent = 'True' ORDER BY name LIMIT ? OFFSET ?"
            page_ids = [r[0] for r in conn.execute(id_sql, [f"%{q}%", per_page, offset]).fetchall()]
            rows = _fetch_company_rows(conn, page_ids) if page_ids else []
        else:
            return {"companies": [], "total": 0, "page": page, "per_page": per_page}

        companies = []
        for r in rows:
            obj_text = r.get("objective", "") or ""
            nace = classify_nace(obj_text)
            cap_amt = r.get("capitalAmount")
            cap_cur = r.get("capitalCurrency", "")
            companies.append({
                "companyId": r["companyId"],
                "name": r["name"],
                "objective": obj_text[:500],
                "address": r.get("fullAddress", "") or "",
                "city": r.get("zipAndPlace", "") or "",
                "register_number": r.get("nativeReferenceNumber", "") or "",
                "court": r.get("courtName", "") or "",
                "founded": r.get("foundedDate", "") or "",
                "dissolved": r.get("dissolutionDate", "") or "",
                "capital": f"{cap_amt} {cap_cur}" if cap_amt else "",
                "nace_codes": nace,
            })

        return {"companies": companies, "total": total, "page": page, "per_page": per_page}
    finally:
        conn.close()


def _de_search_flat(q, page, per_page, offset):
    """Fallback: search using flat openregister.db."""
    conn = _get_flat_db()
    try:
        where, params = [], []
        if q:
            where.append("c.name LIKE ?")
            params.append(f"%{q}%")
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        total = conn.execute(f"SELECT COUNT(*) FROM company c {where_sql}", params).fetchone()[0]
        rows = conn.execute(f"SELECT c.* FROM company c {where_sql} ORDER BY c.name LIMIT ? OFFSET ?",
                            params + [per_page, offset]).fetchall()
        companies = [{k: r[k] for k in r.keys()} for r in rows]
        return {"companies": companies, "total": total, "page": page, "per_page": per_page, "flat": True}
    finally:
        conn.close()


def de_company_detail(company_id):
    """Get full company detail from normalized DB."""
    if not os.path.exists(DB_NORM):
        return None
    conn = _get_norm_db()
    try:
        # Get all names
        names = conn.execute("SELECT * FROM Names WHERE companyId = ? ORDER BY isCurrent DESC, validFrom DESC",
                             (company_id,)).fetchall()
        if not names:
            return None
        current_name = next((n["name"] for n in names if n["isCurrent"] == "True"), names[0]["name"])

        # Company dates
        comp = conn.execute("SELECT * FROM Companies WHERE companyId = ?", (company_id,)).fetchone()

        # Addresses
        addresses = conn.execute("SELECT * FROM Addresses WHERE companyId = ? ORDER BY isCurrent DESC, validFrom DESC",
                                 (company_id,)).fetchall()

        # Objectives
        objectives = conn.execute("SELECT * FROM Objectives WHERE companyId = ? ORDER BY validFrom DESC",
                                  (company_id,)).fetchall()

        # Positions (officers)
        positions = conn.execute("SELECT * FROM Positions WHERE companyId = ? ORDER BY startDate DESC",
                                 (company_id,)).fetchall()

        # Reference numbers
        refs = conn.execute("SELECT * FROM ReferenceNumbers WHERE companyId = ?", (company_id,)).fetchall()

        # Capital
        capital = conn.execute("SELECT * FROM Capital WHERE companyId = ? ORDER BY isCurrent DESC, validFrom DESC",
                               (company_id,)).fetchall()

        # NACE from current objective
        current_obj = next((o["objective"] for o in objectives), "")
        nace = classify_nace(current_obj)

        return {
            "companyId": company_id,
            "name": current_name,
            "nace_codes": nace,
            "company": {k: comp[k] for k in comp.keys()} if comp else {},
            "names": [{k: n[k] for k in n.keys()} for n in names],
            "addresses": [{k: a[k] for k in a.keys()} for a in addresses],
            "objectives": [{k: o[k] for k in o.keys()} for o in objectives],
            "officers": [{k: p[k] for k in p.keys()} for p in positions],
            "references": [{k: r[k] for k in r.keys()} for r in refs],
            "capital": [{k: c[k] for k in c.keys()} for c in capital],
        }
    finally:
        conn.close()


# =====================================================================
#  AUSTRIA — Firmenbuch SOAP API
# =====================================================================

def _soap_request(api_key, body_xml):
    envelope = f"""<?xml version="1.0" encoding="UTF-8"?>
<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope">
  <soap:Header/>
  <soap:Body>
    {body_xml}
  </soap:Body>
</soap:Envelope>"""
    headers = {"Content-Type": "application/soap+xml; charset=utf-8", "X-API-KEY": api_key}
    resp = requests.post(SOAP_URL, data=envelope.encode("utf-8"), headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.text


def at_search(api_key, q, rechtsform="", gericht="", exact=False):
    body = f"""<suc:SUCHEFIRMAREQUEST xmlns:suc="ns://firmenbuch.justiz.gv.at/Abfrage/SucheFirmaRequest">
      <suc:FIRMENWORTLAUT>{_xml_escape(q)}</suc:FIRMENWORTLAUT>
      <suc:EXAKTESUCHE>{"true" if exact else "false"}</suc:EXAKTESUCHE>
      <suc:SUCHBEREICH>1</suc:SUCHBEREICH>
      <suc:GERICHT>{_xml_escape(gericht)}</suc:GERICHT>
      <suc:RECHTSFORM>{_xml_escape(rechtsform)}</suc:RECHTSFORM>
      <suc:RECHTSEIGENSCHAFT/><suc:ORTNR/>
    </suc:SUCHEFIRMAREQUEST>"""
    return _parse_soap_response(_soap_request(api_key, body))


def at_company_detail(api_key, firma_id):
    body = f"""<aus:AUSZUGREQUEST xmlns:aus="ns://firmenbuch.justiz.gv.at/Abfrage/v2/AuszugRequest">
      <aus:FIRMA_ID>{_xml_escape(firma_id)}</aus:FIRMA_ID>
      <aus:VARIANTE>1</aus:VARIANTE>
      <aus:SPRACHE>DE</aus:SPRACHE>
    </aus:AUSZUGREQUEST>"""
    return _parse_soap_response(_soap_request(api_key, body))


def at_changes(api_key, firma_id):
    body = f"""<verf:VERAENDERUNGENFIRMAREQUEST xmlns:verf="ns://firmenbuch.justiz.gv.at/Abfrage/VeraenderungenFirmaRequest">
      <verf:FIRMA_ID>{_xml_escape(firma_id)}</verf:FIRMA_ID>
    </verf:VERAENDERUNGENFIRMAREQUEST>"""
    return _parse_soap_response(_soap_request(api_key, body))


def _xml_escape(text):
    if not text: return ""
    return text.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

def _strip_ns(tag):
    return tag.split("}",1)[1] if "}" in tag else tag

def _elem_to_dict(elem):
    result = {}
    for child in elem:
        key = _strip_ns(child.tag)
        val = _elem_to_dict(child) if list(child) else (child.text or "").strip()
        existing = result.get(key)
        if existing is not None:
            if not isinstance(existing, list): result[key] = [existing]
            result[key].append(val)
        else:
            result[key] = val
    return result

def _parse_soap_response(xml_text):
    root = ET.fromstring(xml_text)
    body = None
    for elem in root.iter():
        if _strip_ns(elem.tag) == "Body": body = elem; break
    if body is None: return {"error": "No SOAP Body"}
    for elem in body.iter():
        if _strip_ns(elem.tag) == "Fault": return {"error": str(_elem_to_dict(elem))}
    return _elem_to_dict(body)


# =====================================================================
#  Flask Routes
# =====================================================================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/search")
def search():
    country = request.args.get("country", "de")
    q = request.args.get("q", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 30))

    if country == "de":
        nace_code = request.args.get("nace", "").strip()
        try:
            return jsonify(de_search(q, page, per_page, nace_code))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    elif country == "at":
        api_key = request.args.get("api_key", "").strip()
        if not api_key:
            return jsonify({"error": "Austrian API key (X-API-KEY) required."}), 400
        rechtsform = request.args.get("rechtsform", "")
        exact = request.args.get("exact", "false").lower() == "true"
        try:
            return jsonify(at_search(api_key, q, rechtsform, exact=exact))
        except requests.HTTPError as e:
            code = e.response.status_code if e.response else 502
            return jsonify({"error": "Invalid API key." if code == 401 else f"API error ({code})"}), code
        except Exception as e:
            return jsonify({"error": str(e)}), 502

    return jsonify({"error": f"Unknown country: {country}"}), 400


@app.route("/api/company/de/<path:company_id>")
def de_detail(company_id):
    try:
        result = de_company_detail(company_id)
        if result is None:
            return jsonify({"error": "Company not found."}), 404
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/company/at/<firma_id>")
def at_detail(firma_id):
    api_key = request.args.get("api_key", "").strip()
    if not api_key: return jsonify({"error": "API key required."}), 400
    try:
        return jsonify(at_company_detail(api_key, firma_id))
    except requests.HTTPError as e:
        return jsonify({"error": f"API error ({e.response.status_code if e.response else 502})"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/company/at/<firma_id>/changes")
def at_company_changes(firma_id):
    api_key = request.args.get("api_key", "").strip()
    if not api_key: return jsonify({"error": "API key required."}), 400
    try:
        return jsonify(at_changes(api_key, firma_id))
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/nace")
def nace_list():
    """Return all NACE codes with labels."""
    return jsonify({code: label for code, label in sorted(NACE_LABELS.items())})


@app.route("/api/export")
def export_xlsx():
    country = request.args.get("country", "de")
    q = request.args.get("q", "").strip()
    nace_code = request.args.get("nace", "").strip()

    wb = Workbook()
    ws = wb.active
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    if country == "de":
        ws.title = "German Companies"
        headers = ["Name","Register Number","Court","Address","City",
                   "Founded","Dissolved","Capital","Business Purpose (Unternehmensgegenstand)",
                   "NACE Code","NACE Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = hfill

        for pg in range(1, 11):
            try:
                result = de_search(q, pg, 500, nace_code)
                cs = result.get("companies", [])
                if not cs: break
                for c in cs:
                    nace = c.get("nace_codes", [])
                    ws.append([
                        c.get("name"), c.get("register_number"), c.get("court"),
                        c.get("address"), c.get("city"),
                        c.get("founded"), c.get("dissolved"), c.get("capital"),
                        c.get("objective",""),
                        ", ".join(n["code"] for n in nace),
                        ", ".join(n["label"] for n in nace),
                    ])
            except Exception: break

    elif country == "at":
        ws.title = "Austrian Companies"
        api_key = request.args.get("api_key", "").strip()
        headers = ["Field", "Value"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = hfill
        if api_key:
            try:
                result = at_search(api_key, q)
                _flatten_to_sheet(ws, result, row=2)
            except Exception as e:
                ws.append([f"Error: {e}"])

    for col in ws.columns:
        mx = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 60)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=companies_{country}.xlsx"})


def _flatten_to_sheet(ws, data, row=2, prefix=""):
    if isinstance(data, dict):
        for k, v in data.items():
            lbl = f"{prefix}{k}" if not prefix else f"{prefix} > {k}"
            if isinstance(v, dict): row = _flatten_to_sheet(ws, v, row, lbl)
            elif isinstance(v, list):
                for i, item in enumerate(v): row = _flatten_to_sheet(ws, item, row, f"{lbl}[{i}]")
            else:
                ws.cell(row=row, column=1, value=lbl)
                ws.cell(row=row, column=2, value=str(v) if v else "")
                row += 1
    return row


@app.route("/api/db-status")
def db_status():
    norm_exists = os.path.exists(DB_NORM)
    flat_exists = os.path.exists(DB_FLAT)
    norm_size = os.path.getsize(DB_NORM) if norm_exists else 0
    flat_size = os.path.getsize(DB_FLAT) if flat_exists else 0
    return jsonify({
        "norm_exists": norm_exists, "norm_size_mb": round(norm_size / 1024 / 1024, 1),
        "flat_exists": flat_exists, "flat_size_mb": round(flat_size / 1024 / 1024, 1),
        "has_objectives": norm_exists,
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, port=port)
