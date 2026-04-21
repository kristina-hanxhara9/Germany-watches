#!/usr/bin/env python3
"""
Classify German companies from openregister.db and handelsregister.db
into industry channels and export to Excel with revenue estimates.
"""

import os
import re
import sqlite3
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DB_NORM = os.path.join(BASE, "handelsregister.db")   # normalized, with objectives
DB_FLAT = os.path.join(BASE, "openregister.db")       # flat, 5.3M companies
OUTPUT  = os.path.join(BASE, "germany_channels.xlsx")

# ── Channel Definitions ─────────────────────────────────────────────────
CHANNELS = {
    "watches_jewellery": {
        "label": "Watches & Jewellery",
        "objective_keywords": [
            "schmuck", "juwelier", "uhren", "goldschmied", "uhrenmacher",
            "edelstein", "uhrmacher", "schmuckwaren", "edelmetall",
            "juwelen", "silberschmied", "bijouterie", "schmuckhandel",
            "uhrenfachgeschäft", "trauringe",
        ],
        "name_keywords": [
            "schmuck", "juwelier", "uhren", "goldschmied", "uhrenmacher",
            "edelstein", "juwelen", "silber", "bijou",
        ],
        "exclude_keywords": ["schmucklos"],
        "revenue_multiplier": 0.8,
    },
    "it_resellers": {
        "label": "IT Resellers",
        "objective_keywords": [
            "it-handel", "systemhaus", "edv-handel", "hardware-handel",
            "computerhandel", "edv-vertrieb", "it-distribution",
            "vertrieb von hardware", "handel mit computern",
            "handel mit edv", "verkauf von it", "it-großhandel",
        ],
        "name_keywords": [
            "it-handel", "systemhaus", "edv-handel", "hardware-handel",
            "computerhandel", "edv-vertrieb", "it-distribution",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 2.0,
    },
    "it_specialists": {
        "label": "IT Specialists",
        "objective_keywords": [
            "softwareentwicklung", "programmierung", "it-beratung",
            "edv-beratung", "it-dienstleistung", "webentwicklung",
            "it-service", "it-consulting", "app-entwicklung",
            "datenverarbeitung", "digitalisierung", "informatik",
            "softwarelösung", "it-sicherheit", "cloud",
        ],
        "name_keywords": [
            "software", "informatik", "it-dienstleistung", "programmierung",
            "it-beratung", "edv-beratung", "webentwicklung", "it-service",
            "it-consulting", "digital",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 1.0,
    },
    "kitchen_specialists": {
        "label": "Kitchen Specialists",
        "objective_keywords": [
            "küchenstudio", "einbauküchen", "küchenhandel",
            "küchenmöbel", "küchenplanung", "kücheneinrichtung",
            "küchencenter", "küchenhaus", "küchengeräte",
            "küchenverkauf", "küchenfachhandel",
        ],
        "name_keywords": [
            "küche", "küchenstudio", "einbauküche", "küchenhandel",
            "küchenmöbel", "küchencenter", "küchenhaus",
        ],
        "exclude_keywords": ["großküche", "industrieküche", "küchenhilfe", "küchenchef"],
        "revenue_multiplier": 1.2,
    },
    "builders_merchants": {
        "label": "Builders Merchants",
        "objective_keywords": [
            "baustoff", "baustoffe", "baustoffhandel", "baubedarf",
            "baumaterial", "baustoffzentrum", "baustoffgroßhandel",
            "handel mit baustoffen", "baustoffmarkt", "baustoffvertrieb",
        ],
        "name_keywords": [
            "baustoff", "baustoffe", "baustoffhandel", "baubedarf",
            "baumaterial", "baustoffzentrum", "baustoffmarkt",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 1.5,
    },
    "plumbing_merchants": {
        "label": "Plumbing Merchants",
        "objective_keywords": [
            "sanitär", "sanitärhandel", "heizung", "klempner",
            "installateur", "sanitärbedarf", "sanitärgroßhandel",
            "haustechnik", "badezimmer", "sanitärtechnik",
            "rohrleitung", "handel mit sanitär", "shk",
            "sanitärfachhandel", "heizungstechnik", "bäderstudio",
            "armaturen", "sanitärausstattung",
        ],
        "name_keywords": [
            "sanitär", "sanitärhandel", "klempner", "installateur",
            "haustechnik", "sanitärbedarf", "heizung sanitär",
            "shk", "bäderstudio", "sanitärtechnik",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 1.3,
    },
    "bookstores": {
        "label": "Bookstores",
        "objective_keywords": [
            "buchhandlung", "buchhandel", "buchladen", "antiquariat",
            "buchversand", "sortimentsbuchhandel", "buchvertrieb",
            "handel mit büchern", "buchgeschäft",
        ],
        "name_keywords": [
            "buchhandlung", "buchhandel", "buchladen", "antiquariat",
            "buchversand", "bücher",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 0.6,
    },
    "diy": {
        "label": "DIY",
        "objective_keywords": [
            "baumarkt", "heimwerker", "bastelbedarf", "gartencenter",
            "heimwerkerbedarf", "heimwerkermarkt", "do-it-yourself",
            "gartenbedarf", "werkzeughandel", "gartenmarkt",
        ],
        "name_keywords": [
            "baumarkt", "heimwerker", "bastelbedarf", "gartencenter",
            "heimwerkerbedarf", "heimwerkermarkt",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 3.0,
    },
    "department_stores": {
        "label": "Department Stores",
        "objective_keywords": [
            "kaufhaus", "warenhaus", "vollsortiment",
            "warensortiment", "sortimentshandel",
        ],
        "name_keywords": [
            "kaufhaus", "warenhaus", "kaufhof", "karstadt",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 5.0,
    },
    "tyre_specialists": {
        "label": "Tyre Specialists",
        "objective_keywords": [
            "reifen", "reifenhandel", "reifenservice", "reifenmontage",
            "vulkanisier", "reifenfachhandel", "reifengroßhandel",
            "handel mit reifen", "felgen", "reifenwechsel",
            "reifenlager", "autoreifen", "reifencenter",
        ],
        "name_keywords": [
            "reifen", "reifenhandel", "reifenservice", "reifencenter",
            "reifenmontage", "vulkanisier", "reifenfachhandel",
        ],
        "exclude_keywords": ["reifensteuer"],
        "revenue_multiplier": 1.2,
    },
    "office_equipment": {
        "label": "Office Equipment Specialists",
        "objective_keywords": [
            "büromaschinen", "bürotechnik", "bürobedarf", "büroeinrichtung",
            "büromöbel", "büroausstattung", "kopierer", "drucker",
            "bürogeräte", "bürofachhandel", "bürohandel",
            "handel mit büro", "bürokommunikation", "bürosysteme",
            "büroorganisation",
        ],
        "name_keywords": [
            "büromaschinen", "bürotechnik", "bürobedarf", "büroeinrichtung",
            "büromöbel", "büroausstattung", "bürofachhandel",
            "bürokommunikation", "bürosysteme", "office",
        ],
        "exclude_keywords": ["office management", "backoffice"],
        "revenue_multiplier": 1.0,
    },
    "stationers": {
        "label": "Stationers",
        "objective_keywords": [
            "schreibwaren", "papeterie", "schreibwarenhandel",
            "papier und schreibwaren", "schulartikel", "bastelbedarf",
            "schreibgerät", "schreibwarengeschäft", "schreibwarenfachhandel",
            "handel mit schreibwaren", "papierwarenhandel", "papierwaren",
        ],
        "name_keywords": [
            "schreibwaren", "papeterie", "schreibwarenhandel",
            "papierwaren", "papierwarenhandel",
        ],
        "exclude_keywords": [],
        "revenue_multiplier": 0.5,
    },
}

# ── Tier-1 cities for geography factor ──────────────────────────────────
TIER1_CITIES = {
    "berlin", "hamburg", "münchen", "munich", "frankfurt", "düsseldorf",
    "stuttgart", "köln", "cologne", "dortmund", "essen", "nürnberg",
    "hannover", "bremen", "dresden", "leipzig",
}

# ── Market context per channel (2025/2026 data) ────────────────────────
# Sources: Statista, Destatis, Mordor Intelligence, IBISWorld, Börsenverein
MARKET_CONTEXT = {
    "watches_jewellery": {
        "market_size_eur": "EUR 6.4B (watches EUR 3.6B + jewellery EUR 2.8B)",
        "market_size_numeric": 6_400_000_000,
        "trend": "stable",
        "trend_factor": 1.0,
        "notes": "Luxury watches 69% of market. Modest growth. Consumer caution on discretionary spend.",
    },
    "it_resellers": {
        "market_size_eur": "EUR 38B (IT distribution & reselling)",
        "market_size_numeric": 38_000_000_000,
        "trend": "growing",
        "trend_factor": 1.15,
        "notes": "Strong demand from AI, cloud migration, cybersecurity. Consolidation ongoing.",
    },
    "it_specialists": {
        "market_size_eur": "EUR 50B+ (IT services & software)",
        "market_size_numeric": 50_000_000_000,
        "trend": "strong_growth",
        "trend_factor": 1.25,
        "notes": "Digitalisierung, AI adoption, skills shortage driving rates up. Fastest growing segment.",
    },
    "kitchen_specialists": {
        "market_size_eur": "EUR 5.3B (kitchen furniture manufacturing)",
        "market_size_numeric": 5_300_000_000,
        "trend": "stable",
        "trend_factor": 1.0,
        "notes": "Small German dwellings (avg 93m²) drive kitchen optimization. Stable demand.",
    },
    "builders_merchants": {
        "market_size_eur": "EUR 22B (Baustoffhandel)",
        "market_size_numeric": 22_000_000_000,
        "trend": "declining",
        "trend_factor": 0.90,
        "notes": "Construction sector downturn since 2023. Housing starts down 25%. Recovery expected 2026+.",
    },
    "plumbing_merchants": {
        "market_size_eur": "EUR 4.75B (SHK wholesale, total SHK-Handwerk EUR 58B)",
        "market_size_numeric": 4_750_000_000,
        "trend": "growing",
        "trend_factor": 1.10,
        "notes": "Heat pump mandate (65% renewable in new heating). SHK sector booming despite construction dip.",
    },
    "bookstores": {
        "market_size_eur": "EUR 9.9B (total Buchhandel gross turnover 2024)",
        "market_size_numeric": 9_900_000_000,
        "trend": "stable",
        "trend_factor": 0.95,
        "notes": "Stationary retail EUR 4.1B (41.3%). +1.8% in 2024. BookTok trend helps. Fixed book pricing law protects margins.",
    },
    "diy": {
        "market_size_eur": "EUR 59.2B (DIY & hardware store 2025)",
        "market_size_numeric": 59_200_000_000,
        "trend": "recovering",
        "trend_factor": 1.05,
        "notes": "Eco-friendly/sustainable products trending. Bauhaus, Obi, Hornbach dominate. Garden segment soft.",
    },
    "department_stores": {
        "market_size_eur": "EUR 12B (Warenhäuser/Kaufhäuser)",
        "market_size_numeric": 12_000_000_000,
        "trend": "declining",
        "trend_factor": 0.80,
        "notes": "Structural decline. Galeria Karstadt Kaufhof repeated insolvency. E-commerce cannibalization.",
    },
    "tyre_specialists": {
        "market_size_eur": "EUR 8B (Reifenhandel & service)",
        "market_size_numeric": 8_000_000_000,
        "trend": "stable",
        "trend_factor": 1.0,
        "notes": "Seasonal business (winter/summer swap). EV tyre segment growing. Price pressure from online.",
    },
    "office_equipment": {
        "market_size_eur": "EUR 7B (Bürotechnik & Bürobedarf)",
        "market_size_numeric": 7_000_000_000,
        "trend": "growing",
        "trend_factor": 1.10,
        "notes": "Home office trend sustained. AI-powered tools. Print declining but office furniture strong.",
    },
    "stationers": {
        "market_size_eur": "EUR 1.5B (Schreibwaren Einzelhandel)",
        "market_size_numeric": 1_500_000_000,
        "trend": "declining",
        "trend_factor": 0.85,
        "notes": "Digital substitution. Niche survival via premium/craft segments. Back-to-school seasonal peaks.",
    },
}

# ── Revenue estimation ──────────────────────────────────────────────────

def _capital_score(capital_amount):
    """Score 0-4 based on share capital (Stammkapital)."""
    if not capital_amount or capital_amount <= 0:
        return 1  # unknown -> assume small
    if capital_amount < 25000:
        return 0  # micro (UG)
    if capital_amount == 25000:
        return 1  # standard GmbH minimum
    if capital_amount <= 100000:
        return 2
    if capital_amount <= 500000:
        return 3
    return 4

def _register_score(register_type):
    """Score 0-3 based on register type."""
    rt = (register_type or "").upper()
    if rt == "HRB":
        return 2
    if rt == "HRA":
        return 1
    return 0

def _officer_score(officer_count):
    """Score 0-4 based on active officer count."""
    if not officer_count or officer_count <= 0:
        return 0
    if officer_count <= 2:
        return 1
    if officer_count <= 5:
        return 2
    if officer_count <= 10:
        return 3
    return 4

def _geo_score(address_or_city):
    """Score 0-1 based on geographic tier."""
    if not address_or_city:
        return 0
    text = address_or_city.lower()
    for city in TIER1_CITIES:
        if city in text:
            return 1
    return 0

def _age_score(founded_date):
    """Score 0-4 based on company age. Older = more established = likely larger."""
    if not founded_date:
        return 1  # unknown -> assume mid
    try:
        year = int(str(founded_date)[:4])
    except (ValueError, TypeError):
        return 1
    age = 2026 - year
    if age <= 2:
        return 0   # startup
    if age <= 5:
        return 1   # young
    if age <= 15:
        return 2   # established
    if age <= 30:
        return 3   # mature
    return 4        # long-standing (30+ years)

REVENUE_RANGES = [
    "EUR 0 - 500K",
    "EUR 500K - 2M",
    "EUR 2M - 10M",
    "EUR 10M - 50M",
    "EUR 50M+",
]

def _city_tier_label(address_or_city):
    """Return city tier label."""
    if not address_or_city:
        return "Tier 3"
    text = address_or_city.lower()
    for city in TIER1_CITIES:
        if city in text:
            return "Tier 1"
    return "Tier 2/3"


def _trend_label(channel_key):
    """Return market trend label for display."""
    mkt = MARKET_CONTEXT.get(channel_key, {})
    trend = mkt.get("trend", "unknown")
    factor = mkt.get("trend_factor", 1.0)
    pct = round((factor - 1.0) * 100)
    sign = "+" if pct >= 0 else ""
    return f"{trend.replace('_', ' ').title()} ({sign}{pct}%)"


def sector_avg_revenue(channel_key, company_count):
    """Calculate sector average revenue = market size / number of companies."""
    mkt = MARKET_CONTEXT.get(channel_key, {})
    market_size = mkt.get("market_size_numeric", 0)
    if not market_size or not company_count:
        return 0
    return int(market_size / company_count)


def estimate_revenue(capital_amount, register_type, officer_count, address,
                     channel_key, founded_date=None):
    """Estimate revenue range using multi-factor model + market trends."""
    cs = _capital_score(capital_amount)
    rs = _register_score(register_type)
    os_ = _officer_score(officer_count)
    gs = _geo_score(address)
    ag = _age_score(founded_date)

    mkt = MARKET_CONTEXT.get(channel_key, {})
    trend_factor = mkt.get("trend_factor", 1.0)

    # Normalize each to 0-1 range, then weight
    weighted = (
        (cs / 4.0) * 0.20 +
        (rs / 2.0) * 0.10 +
        (os_ / 4.0) * 0.20 +
        (ag / 4.0) * 0.20 +
        gs * 0.10 +
        min(1.0, trend_factor - 0.5) * 0.20
    )

    # Apply industry revenue multiplier from channel definition
    multiplier = CHANNELS.get(channel_key, {}).get("revenue_multiplier", 1.0)
    weighted = min(1.0, weighted * (1 + (multiplier - 1) * 0.3))

    # Map to revenue range
    if weighted < 0.15:
        idx = 0
    elif weighted < 0.30:
        idx = 1
    elif weighted < 0.50:
        idx = 2
    elif weighted < 0.70:
        idx = 3
    else:
        idx = 4

    return REVENUE_RANGES[idx]


# ── Database helpers ────────────────────────────────────────────────────

def open_db(path):
    """Open a SQLite DB with read-only optimizations."""
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA mmap_size=268435456")
    conn.execute("PRAGMA cache_size=-64000")  # 64MB cache
    return conn


def build_fts_query(keywords):
    """Build an FTS5 MATCH expression from keywords."""
    # Quote each keyword and OR them together
    parts = [f'"{kw}"' for kw in keywords]
    return " OR ".join(parts)


def build_like_conditions(column, keywords):
    """Build SQL LIKE conditions for keywords."""
    conditions = [f"lower({column}) LIKE ?" for _ in keywords]
    params = [f"%{kw}%" for kw in keywords]
    return " OR ".join(conditions), params


# ── Main classification logic ───────────────────────────────────────────

def classify_channel_handelsregister(conn_norm, channel_key, channel_def):
    """Classify companies from handelsregister.db using FTS on objectives."""
    print(f"  [handelsregister] Querying objectives FTS...")

    # Phase 1: FTS query on ObjectivesFts
    fts_match = build_fts_query(channel_def["objective_keywords"])
    try:
        rows = conn_norm.execute(
            "SELECT companyId, name FROM ObjectivesFts WHERE ObjectivesFts MATCH ?",
            (fts_match,)
        ).fetchall()
    except Exception as e:
        print(f"    FTS error: {e}, falling back to LIKE")
        rows = []

    company_ids = set()
    for r in rows:
        cid = r["companyId"]
        # Check excludes
        if channel_def.get("exclude_keywords"):
            text = (r["name"] or "").lower()
            if any(ex in text for ex in channel_def["exclude_keywords"]):
                continue
        company_ids.add(cid)

    # Phase 2: Supplement with LIKE on Objectives table for compound words
    like_cond, like_params = build_like_conditions("objective", channel_def["objective_keywords"])
    try:
        like_rows = conn_norm.execute(
            f"SELECT DISTINCT companyId FROM Objectives WHERE ({like_cond})",
            like_params
        ).fetchall()
        for r in like_rows:
            company_ids.add(r["companyId"])
    except Exception as e:
        print(f"    LIKE fallback error: {e}")

    print(f"  [handelsregister] Found {len(company_ids)} matching companyIds")

    if not company_ids:
        return []

    # Phase 3: Fetch full details in batches
    results = []
    id_list = list(company_ids)
    batch_size = 500

    for i in range(0, len(id_list), batch_size):
        batch = id_list[i:i + batch_size]
        placeholders = ",".join("?" * len(batch))

        # Names (current)
        names = {}
        for r in conn_norm.execute(
            f"SELECT companyId, name FROM Names WHERE isCurrent='True' AND companyId IN ({placeholders})",
            batch
        ):
            names[r["companyId"]] = r["name"]

        # Addresses (current)
        addresses = {}
        for r in conn_norm.execute(
            f"SELECT companyId, fullAddress, zipAndPlace, zipCode FROM Addresses WHERE isCurrent='True' AND companyId IN ({placeholders})",
            batch
        ):
            addresses[r["companyId"]] = {
                "fullAddress": r["fullAddress"],
                "zipAndPlace": r["zipAndPlace"],
                "zipCode": r["zipCode"],
            }

        # Capital (current)
        capitals = {}
        for r in conn_norm.execute(
            f"SELECT companyId, capitalAmount, capitalCurrency FROM Capital WHERE isCurrent='True' AND companyId IN ({placeholders})",
            batch
        ):
            capitals[r["companyId"]] = r["capitalAmount"]

        # Officer count
        officer_counts = {}
        for r in conn_norm.execute(
            f"SELECT companyId, COUNT(*) as cnt FROM Positions WHERE endDate IS NULL AND companyId IN ({placeholders}) GROUP BY companyId",
            batch
        ):
            officer_counts[r["companyId"]] = r["cnt"]

        # Objectives text (current, for display)
        objectives = {}
        for r in conn_norm.execute(
            f"SELECT companyId, objective FROM Objectives WHERE companyId IN ({placeholders})",
            batch
        ):
            objectives[r["companyId"]] = r["objective"]

        # Founding date
        founded_dates = {}
        for r in conn_norm.execute(
            f"SELECT companyId, foundedDate FROM Companies WHERE companyId IN ({placeholders})",
            batch
        ):
            founded_dates[r["companyId"]] = r["foundedDate"]

        for cid in batch:
            name = names.get(cid, "")
            addr = addresses.get(cid, {})
            cap = capitals.get(cid)
            oc = officer_counts.get(cid, 0)
            obj = objectives.get(cid, "")
            founded = founded_dates.get(cid)

            # Extract city from zipAndPlace
            zp = addr.get("zipAndPlace", "") or ""
            city = zp.split(" ", 1)[1] if " " in zp else zp

            # Extract register type from companyId (e.g., B1102_HRB2395)
            reg_type = ""
            m = re.search(r"_(HRB|HRA|VR|GnR|PR)", cid)
            if m:
                reg_type = m.group(1)

            full_addr = addr.get("fullAddress", "") or city
            results.append({
                "company_id": cid,
                "name": name,
                "address": addr.get("fullAddress", ""),
                "city": city,
                "federal_state": "",
                "register_type": reg_type,
                "status": "active",
                "founded": founded or "",
                "officer_count": oc,
                "capital": cap,
                "city_tier": _city_tier_label(full_addr),
                "market_trend": _trend_label(channel_key),
                "objective": (obj or "")[:500],
                "revenue_estimate": estimate_revenue(
                    cap, reg_type, oc, full_addr,
                    channel_key, founded,
                ),
                "source": "handelsregister",
            })

    return results


def normalize_id_segments(compound_id):
    """Extract all simple ID segments from a compound handelsregister ID."""
    # "B1102_HRB2395-F1103_HRB102428B" -> {"B1102_HRB2395", "F1103_HRB102428B"}
    segments = set()
    for seg in compound_id.split("-"):
        segments.add(seg)
        # Also add version without trailing letter on register number
        # e.g., F1103_HRB102428B -> F1103_HRB102428
        cleaned = re.sub(r"[A-Za-z]+$", "", seg)
        if cleaned != seg:
            segments.add(cleaned)
    return segments


def classify_channel_openregister(conn_flat, channel_key, channel_def, seen_ids):
    """Classify companies from openregister.db using name LIKE, excluding already-seen IDs."""
    print(f"  [openregister] Querying company names...")

    like_cond, like_params = build_like_conditions("c.name", channel_def["name_keywords"])

    query = f"""
        SELECT c.company_number, c.name, c.registered_address, c.federal_state,
               c.register_art, c.current_status, c.registered_office
        FROM company c
        WHERE c.current_status = 'currently registered'
          AND ({like_cond})
    """

    rows = conn_flat.execute(query, like_params).fetchall()
    print(f"  [openregister] Found {len(rows)} name matches")

    # Filter out companies already found in handelsregister
    new_rows = []
    for r in rows:
        cn = r["company_number"]
        if cn in seen_ids:
            continue
        # Check excludes
        if channel_def.get("exclude_keywords"):
            text = (r["name"] or "").lower()
            if any(ex in text for ex in channel_def["exclude_keywords"]):
                continue
        new_rows.append(r)

    print(f"  [openregister] {len(new_rows)} new companies (after dedup)")

    if not new_rows:
        return []

    # Fetch officer counts in batches
    results = []
    batch_size = 500
    for i in range(0, len(new_rows), batch_size):
        batch = new_rows[i:i + batch_size]
        cids = [r["company_number"] for r in batch]
        placeholders = ",".join("?" * len(cids))

        officer_counts = {}
        try:
            for oc_row in conn_flat.execute(
                f"SELECT company_id, COUNT(*) as cnt FROM officer WHERE company_id IN ({placeholders}) GROUP BY company_id",
                cids
            ):
                officer_counts[oc_row["company_id"]] = oc_row["cnt"]
        except Exception:
            pass

        for r in batch:
            cn = r["company_number"]
            oc = officer_counts.get(cn, 0)
            address = r["registered_address"] or ""
            city = r["registered_office"] or ""

            full_addr = address or city
            results.append({
                "company_id": cn,
                "name": r["name"],
                "address": address,
                "city": city,
                "federal_state": r["federal_state"] or "",
                "register_type": r["register_art"] or "",
                "status": r["current_status"] or "",
                "founded": "",
                "officer_count": oc,
                "capital": None,
                "city_tier": _city_tier_label(full_addr),
                "market_trend": _trend_label(channel_key),
                "objective": "",
                "revenue_estimate": estimate_revenue(
                    None, r["register_art"], oc,
                    full_addr, channel_key, None,
                ),
                "source": "openregister",
            })

    return results


def classify_all_channels():
    """Run classification for all channels across both databases."""
    print("Opening databases...")
    conn_norm = open_db(DB_NORM) if os.path.exists(DB_NORM) else None
    conn_flat = open_db(DB_FLAT) if os.path.exists(DB_FLAT) else None

    if not conn_norm and not conn_flat:
        print("ERROR: No databases found!")
        sys.exit(1)

    all_results = {}

    for channel_key, channel_def in CHANNELS.items():
        print(f"\n{'='*60}")
        print(f"Channel: {channel_def['label']}")
        print(f"{'='*60}")

        channel_results = []
        seen_ids = set()

        # Phase 1: handelsregister (richer data)
        if conn_norm:
            hr_results = classify_channel_handelsregister(conn_norm, channel_key, channel_def)
            channel_results.extend(hr_results)
            # Build set of all ID segments for dedup
            for r in hr_results:
                segments = normalize_id_segments(r["company_id"])
                seen_ids.update(segments)

        # Phase 2: openregister (supplement)
        if conn_flat:
            or_results = classify_channel_openregister(conn_flat, channel_key, channel_def, seen_ids)
            channel_results.extend(or_results)

        all_results[channel_key] = channel_results
        print(f"  TOTAL: {len(channel_results)} companies")

    if conn_norm:
        conn_norm.close()
    if conn_flat:
        conn_flat.close()

    return all_results


# ── Excel export ────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

COLUMNS = [
    ("Company Name", 40),
    ("Address", 45),
    ("City", 20),
    ("Federal State", 18),
    ("Register Type", 13),
    ("Status", 15),
    ("Founded", 12),
    ("Company Age (yrs)", 14),
    ("Officers", 10),
    ("Share Capital (EUR)", 18),
    ("City Tier", 10),
    ("Market Trend", 14),
    ("Sector Avg Rev (EUR)", 20),
    ("Est. Revenue Range", 18),
    ("Business Purpose", 60),
    ("Data Source", 14),
]


def write_excel(all_results, output_path):
    """Write classified results to Excel workbook."""
    print(f"\nWriting Excel to {output_path}...")
    wb = Workbook()

    # ── Summary sheet ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = [
        "Channel", "Companies", "With Capital Data", "Avg Capital (EUR)",
        "Top Federal State", "Top City", "Market Size (EUR)",
        "Sector Avg Rev (EUR)", "Market Trend", "Trend Factor", "Market Notes",
    ]
    for col, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Trend colors
    trend_fills = {
        "strong_growth": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "growing":       PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid"),
        "recovering":    PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid"),
        "stable":        PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        "declining":     PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid"),
    }

    row = 2
    for channel_key, channel_def in CHANNELS.items():
        results = all_results.get(channel_key, [])
        mkt = MARKET_CONTEXT.get(channel_key, {})

        capitals = [r["capital"] for r in results if r.get("capital") and r["capital"] > 0]
        avg_cap = int(sum(capitals) / len(capitals)) if capitals else 0

        states = Counter(r["federal_state"] for r in results if r["federal_state"])
        top_state = states.most_common(1)[0][0] if states else "N/A"

        cities = Counter(r["city"] for r in results if r["city"])
        top_city = cities.most_common(1)[0][0] if cities else "N/A"

        sect_avg = sector_avg_revenue(channel_key, len(results)) if results else 0

        ws_sum.cell(row=row, column=1, value=channel_def["label"])
        ws_sum.cell(row=row, column=2, value=len(results))
        ws_sum.cell(row=row, column=3, value=len(capitals))
        c4 = ws_sum.cell(row=row, column=4, value=avg_cap)
        c4.number_format = '#,##0'
        ws_sum.cell(row=row, column=5, value=top_state)
        ws_sum.cell(row=row, column=6, value=top_city)
        ws_sum.cell(row=row, column=7, value=mkt.get("market_size_eur", "N/A"))
        c8 = ws_sum.cell(row=row, column=8, value=sect_avg)
        c8.number_format = '#,##0'
        trend = mkt.get("trend", "N/A")
        trend_cell = ws_sum.cell(row=row, column=9, value=trend.replace("_", " ").title())
        if trend in trend_fills:
            trend_cell.fill = trend_fills[trend]
        ws_sum.cell(row=row, column=10, value=mkt.get("trend_factor", ""))
        ws_sum.cell(row=row, column=11, value=mkt.get("notes", ""))
        row += 1

    # Auto-width summary
    col_widths = [28, 12, 16, 18, 22, 18, 42, 20, 16, 12, 70]
    for col, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = w
    ws_sum.freeze_panes = "A2"

    # ── Channel sheets ──
    for channel_key, channel_def in CHANNELS.items():
        results = all_results.get(channel_key, [])
        label = channel_def["label"]
        sheet_name = label[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Sector average revenue (market size / companies in sector)
        sect_avg = sector_avg_revenue(channel_key, len(results)) if results else 0

        # Headers
        for col, (header, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

        # Data rows
        for i, r in enumerate(results):
            row_num = i + 2
            c = 1
            # Company Name
            ws.cell(row=row_num, column=c, value=r["name"]); c += 1
            # Address
            ws.cell(row=row_num, column=c, value=r["address"]); c += 1
            # City
            ws.cell(row=row_num, column=c, value=r["city"]); c += 1
            # Federal State
            ws.cell(row=row_num, column=c, value=r["federal_state"]); c += 1
            # Register Type
            ws.cell(row=row_num, column=c, value=r["register_type"]); c += 1
            # Status
            ws.cell(row=row_num, column=c, value=r["status"]); c += 1

            # Founded
            founded = r.get("founded", "")
            ws.cell(row=row_num, column=c, value=founded); c += 1

            # Company Age
            age_val = ""
            if founded:
                try:
                    age_val = 2026 - int(str(founded)[:4])
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_num, column=c, value=age_val); c += 1

            # Officers
            ws.cell(row=row_num, column=c, value=r["officer_count"]); c += 1

            # Share Capital
            cap_cell = ws.cell(row=row_num, column=c, value=r["capital"])
            if r["capital"]:
                cap_cell.number_format = '#,##0'
            c += 1

            # City Tier (fact)
            ws.cell(row=row_num, column=c, value=r.get("city_tier", "")); c += 1

            # Market Trend (fact)
            ws.cell(row=row_num, column=c, value=r.get("market_trend", "")); c += 1

            # Sector Avg Revenue
            avg_cell = ws.cell(row=row_num, column=c, value=sect_avg)
            avg_cell.number_format = '#,##0'
            c += 1

            # Est. Revenue Range
            ws.cell(row=row_num, column=c, value=r["revenue_estimate"]); c += 1

            # Business Purpose
            ws.cell(row=row_num, column=c, value=r["objective"]); c += 1

            # Data Source
            ws.cell(row=row_num, column=c, value=r["source"]); c += 1

            # Alternating row fill
            if i % 2 == 1:
                for col in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row_num, column=col).fill = ALT_FILL

        print(f"  Sheet '{sheet_name}': {len(results)} rows | Sector avg: EUR {sect_avg:,.0f}")

    wb.save(output_path)
    print(f"\nDone! Saved to: {output_path}")


# ── Main ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    output = OUTPUT
    if len(sys.argv) > 1:
        output = sys.argv[1]

    all_results = classify_all_channels()

    # Print summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    total = 0
    for channel_key, channel_def in CHANNELS.items():
        count = len(all_results.get(channel_key, []))
        total += count
        print(f"  {channel_def['label']:30s} {count:>8,}")
    print(f"  {'TOTAL':30s} {total:>8,}")

    write_excel(all_results, output)
