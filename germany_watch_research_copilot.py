"""
Germany Watch Retailer Research — GitHub Copilot SDK
=====================================================
Applies the watch-retailer definition to every German candidate:

    INCLUDE : sale of watches mandatory AND >=80% turnover from
              watches / fine jewellery / watch-repair.
    EXCLUDE : >50% fashion accessories (Pandora, Thomas Sabo, Modeschmuck)
              OR >50% repair-only with no watch retail
              OR no watch sales at all.
    REVIEW  : genuinely ambiguous after web research.

API notes (matches the real github-copilot-sdk Python docs):
  - `CopilotClient()` is started with `await client.start()`
    and stopped with `await client.stop()`
  - `client.create_session(...)` is awaitable; can be used as an
    async context manager via `async with await client.create_session(...) as s:`
  - Prompt + wait-until-idle in one call: `await session.send_and_wait(prompt)`
  - Event data types are pattern-matched, not string-compared
  - `infinite_sessions={"enabled": False}` gives a fully isolated session per row

Setup:
    pip install github-copilot-sdk pandas openpyxl
    gh auth login          # copilot-sdk uses the gh CLI credentials

Run:
    python germany_watch_research_copilot.py

The agent uses the SDK's built-in `url` fetch tool (no Playwright / no MCP
server needed). Searches go through DuckDuckGo HTML + direct URLs on
gelbeseiten.de, northdata.com, handelsregister.de — all of which serve
plain HTML that the url-fetch tool can read directly.
"""

import asyncio
import json
import logging
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from copilot import CopilotClient
from copilot.session import PermissionRequestResult
from copilot.generated.session_events import (
    PermissionRequest,
    AssistantMessageData,
    AssistantMessageDeltaData,
    SessionIdleData,
)

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_FILE  = "watch_retailers_germany.xlsx"
INPUT_SHEET = "INCLUDE"
OUTPUT_FILE = "watch_retailers_germany_enriched.xlsx"
CHECKPOINT  = "checkpoint_de_watches.json"

COL_NAME      = "name"
COL_ZIP       = "zip"
COL_CITY      = "city"
COL_REGISTER  = "register"
COL_OBJECTIVE = "objective"
COL_ACTIVE    = "active"

MODEL       = "gpt-5"
STREAMING   = True
CONCURRENCY = 3

SKILL_DIR = str(Path(__file__).parent / ".github" / "skills")

WZ_SECTORS = {
    "47.77": "Einzelhandel mit Uhren und Schmuck",
    "47.78": "Einzelhandel mit sonstigen Gütern a.n.g.",
    "46.48": "Großhandel mit Uhren und Schmuck",
    "95.25": "Reparatur von Uhren und Schmuck",
    "32.12": "Herstellung von Schmuck",
    "32.13": "Herstellung von Fantasieschmuck",
}

OUTPUT_FIELDS = [
    "website", "phone_number", "address",
    "google_maps_url", "google_maps_rating", "google_maps_review_count",
    "opening_hours", "about",
    "products_sold", "watch_brands_carried", "jewellery_brands_carried",
    "own_brands", "offers_repair_services",
    "chain_or_group", "parent_company", "number_of_locations",
    "annual_turnover", "employee_count",
    "target_customers", "price_positioning",
    "online_shop_url", "social_media", "recent_news",
    "northdata_url", "handelsregister_url", "gelbeseiten_url",
    "classification", "classification_reason",
    "data_confidence", "sources_checked", "research_error",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("research_de.log")],
)
log = logging.getLogger(__name__)


# ── Permission handler ────────────────────────────────────────────────────────
# request.kind.value is one of: shell, write, read, mcp, custom-tool, url, memory, hook
# We approve read / url / mcp / memory (needed for web research) and deny shell + write.

APPROVED_KINDS = {"read", "url", "mcp", "memory", "custom-tool", "hook"}


def permission_handler(
    request: PermissionRequest,
    invocation: dict,
) -> PermissionRequestResult:
    kind = request.kind.value
    if kind in APPROVED_KINDS:
        return PermissionRequestResult(kind="approved")
    # shell + write are denied for safety
    return PermissionRequestResult(kind="denied-interactively-by-user")


# ── Prompt ────────────────────────────────────────────────────────────────────

def build_prompt(name: str, zip_code: str, city: str, register: str, objective: str) -> str:
    obj = (objective or "").strip().replace("\n", " ")
    if len(obj) > 800:
        obj = obj[:800] + "…"
    return (
        f"Use the /germany-watch-research skill to research this company.\n\n"
        f"Company name         : {name}\n"
        f"ZIP / PLZ            : {zip_code}\n"
        f"City / Stadt         : {city}\n"
        f"Handelsregister Nr.  : {register or '—'}\n"
        f"Country              : Germany\n\n"
        f"Unternehmensgegenstand (Handelsregister objective):\n"
        f"{obj or '—'}\n\n"
        f"Apply the watch-retailer definition in the skill exactly. "
        f"Return only JSON."
    )


# ── JSON parser ───────────────────────────────────────────────────────────────

def parse_json(raw: str) -> dict:
    clean = re.sub(r"```(?:json)?|```", "", raw).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", clean, re.DOTALL)
        if m:
            return json.loads(m.group())
        raise


# ── Research one company ──────────────────────────────────────────────────────

async def research_one(
    client: CopilotClient,
    name: str,
    zip_code: str,
    city: str,
    register: str,
    objective: str,
) -> dict:
    prompt = build_prompt(name, zip_code, city, register, objective)

    for attempt in range(1, 4):
        try:
            final_messages: list[str] = []
            delta_buffer:   list[str] = []

            def on_event(event):
                # Prefer the final (non-delta) AssistantMessageData for reliability;
                # fall back to accumulated deltas if no final message arrives.
                data = event.data
                if isinstance(data, AssistantMessageData):
                    if data.content:
                        final_messages.append(data.content)
                elif isinstance(data, AssistantMessageDeltaData):
                    if data.delta_content:
                        delta_buffer.append(data.delta_content)

            async with await client.create_session(
                on_permission_request=permission_handler,
                model=MODEL,
                streaming=STREAMING,
                skill_directories=[SKILL_DIR],
                infinite_sessions={"enabled": False},
            ) as session:
                session.on(on_event)
                # send_and_wait returns when the session becomes idle
                await asyncio.wait_for(session.send_and_wait(prompt), timeout=240)

            full = "\n".join(final_messages).strip() or "".join(delta_buffer).strip()
            if not full:
                raise ValueError("Empty response from model")

            result = parse_json(full)
            result["research_error"] = None

            log.info(
                "OK  %-35s | %-7s | %s | %d sources",
                name[:35],
                result.get("classification", "?"),
                result.get("data_confidence", "?"),
                len(result.get("sources_checked") or []),
            )
            return result

        except asyncio.TimeoutError:
            log.warning("Timeout: %s (attempt %d/3)", name, attempt)
            await asyncio.sleep(15 * attempt)

        except (json.JSONDecodeError, ValueError) as e:
            log.warning("Parse error: %s — %s", name, e)
            break

        except Exception as e:
            log.warning("Error: %s — %s (attempt %d/3)", name, e, attempt)
            if attempt < 3:
                await asyncio.sleep(15 * attempt)

    empty = {f: None for f in OUTPUT_FIELDS}
    empty["research_error"] = "Failed after 3 attempts"
    empty["data_confidence"] = "low"
    empty["classification"] = "REVIEW"
    empty["classification_reason"] = "Research failed — manual check required"
    log.error("FAIL %s", name)
    return empty


# ── Checkpoint ────────────────────────────────────────────────────────────────

def load_ckpt() -> dict:
    p = Path(CHECKPOINT)
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else {}

def save_ckpt(data: dict) -> None:
    Path(CHECKPOINT).write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ── Excel output ──────────────────────────────────────────────────────────────

def save_excel(df: pd.DataFrame) -> None:
    df.to_excel(OUTPUT_FILE, index=False)
    wb  = load_workbook(OUTPUT_FILE)
    ws  = wb.active

    input_cols = {COL_NAME, COL_ZIP, COL_CITY, COL_REGISTER, COL_OBJECTIVE, COL_ACTIVE}

    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="0C447C")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col = ws.cell(row=1, column=cell.column).value
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col not in input_cols:
                cell.fill = PatternFill("solid", fgColor="E6F1FB")

            if col == "classification" and cell.value:
                v = str(cell.value).upper()
                if v == "INCLUDE":
                    cell.fill = PatternFill("solid", fgColor="D4EDDA")
                elif v == "EXCLUDE":
                    cell.fill = PatternFill("solid", fgColor="F8D7DA")
                elif v == "REVIEW":
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")

            if col == "research_error" and cell.value:
                cell.fill = PatternFill("solid", fgColor="FCEBEB")

    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 50)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(OUTPUT_FILE)
    log.info("Saved -> %s", OUTPUT_FILE)


# ── Pipeline ──────────────────────────────────────────────────────────────────

async def run() -> None:
    p = Path(INPUT_FILE)
    if not p.exists():
        raise FileNotFoundError(
            f"\nInput not found: {p.resolve()}\n"
            f"Run find_watch_retailers.py first, or place '{INPUT_FILE}' next to this script."
        )

    skill_path = Path(SKILL_DIR) / "germany-watch-research"
    if not skill_path.exists():
        raise FileNotFoundError(
            f"Skill directory not found: {skill_path.resolve()}\n"
            f"Make sure .github/skills/germany-watch-research/SKILL.md exists."
        )

    df    = pd.read_excel(p, sheet_name=INPUT_SHEET, dtype=str).fillna("")
    total = len(df)
    ckpt  = load_ckpt()
    start = time.time()

    log.info("=" * 60)
    log.info("Germany Watch Research | %d rows | %d done | model: %s", total, len(ckpt), MODEL)
    log.info("Skill: %s", SKILL_DIR)
    log.info("Context isolation: infinite_sessions disabled per row")
    log.info("=" * 60)

    client = CopilotClient()
    await client.start()
    try:
        sem = asyncio.Semaphore(CONCURRENCY)

        async def process(idx: int) -> None:
            name     = str(df.iloc[idx].get(COL_NAME,      "")).strip()
            zip_code = str(df.iloc[idx].get(COL_ZIP,       "")).strip()
            city     = str(df.iloc[idx].get(COL_CITY,      "")).strip()
            register = str(df.iloc[idx].get(COL_REGISTER,  "")).strip()
            obj      = str(df.iloc[idx].get(COL_OBJECTIVE, "")).strip()
            key      = f"{idx}:{name}:{zip_code}"

            if key in ckpt:
                return

            async with sem:
                result = await research_one(client, name, zip_code, city, register, obj)

            ckpt[key] = result
            save_ckpt(ckpt)

            n         = len(ckpt)
            elapsed   = time.time() - start
            rate      = n / elapsed if elapsed else 0
            remaining = (total - n) / rate if rate else 0
            log.info(
                "[%d/%d] %-35s | ETA %dm %ds",
                n, total, name[:35],
                int(remaining // 60), int(remaining % 60),
            )

        await asyncio.gather(*[process(i) for i in range(total)])
    finally:
        await client.stop()

    # Merge results back
    for f in OUTPUT_FIELDS:
        df[f] = ""
    for idx in range(total):
        name     = str(df.iloc[idx].get(COL_NAME, "")).strip()
        zip_code = str(df.iloc[idx].get(COL_ZIP,  "")).strip()
        result   = ckpt.get(f"{idx}:{name}:{zip_code}", {})
        for f in OUTPUT_FIELDS:
            val = result.get(f, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            if isinstance(val, bool):
                val = "yes" if val else "no"
            df.at[idx, f] = val if val is not None else ""

    save_excel(df)

    elapsed = time.time() - start
    log.info("=" * 60)
    log.info(
        "DONE %.1f min | INCLUDE: %d | EXCLUDE: %d | REVIEW: %d | Err: %d",
        elapsed / 60,
        sum(1 for v in ckpt.values() if (v.get("classification") or "").upper() == "INCLUDE"),
        sum(1 for v in ckpt.values() if (v.get("classification") or "").upper() == "EXCLUDE"),
        sum(1 for v in ckpt.values() if (v.get("classification") or "").upper() == "REVIEW"),
        sum(1 for v in ckpt.values() if v.get("research_error")),
    )
    log.info("=" * 60)


if __name__ == "__main__":
    asyncio.run(run())
